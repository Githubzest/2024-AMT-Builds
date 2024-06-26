import streamlit as st
import numpy as np
import skfuzzy as fuzz
import openpyxl
import io

def compare_excel_files(file1, file2):
    # Load the two Excel files
    workbook1 = openpyxl.load_workbook(io.BytesIO(file1))
    workbook2 = openpyxl.load_workbook(io.BytesIO(file2))
    sheet1 = workbook1.active
    sheet2 = workbook2.active

    # Define the input variable range (cell value differences)
    value_range = np.arange(-100, 101, 1)

    # Define membership functions for similarity
    identical_membership = fuzz.membership.trimf(value_range, [-1, 0, 1])
    different_membership = fuzz.membership.trimf(value_range, [-100, 0, 100])

    # Iterate over cells and compare values
    comparison_results = []
    for row in range(1, max(sheet1.max_row, sheet2.max_row) + 1):
        for col in range(1, max(sheet1.max_column, sheet2.max_column) + 1):
            cell1_value = sheet1.cell(row=row, column=col).value or 0
            cell2_value = sheet2.cell(row=row, column=col).value or 0

            # Convert cell values to numeric if possible
            try:
                cell1_value = float(cell1_value) if cell1_value else 0
                cell2_value = float(cell2_value) if cell2_value else 0
            except ValueError:
                cell1_value = 0
                cell2_value = 0

            value_diff = cell1_value - cell2_value

            # Fuzzy inference
            identical_level = fuzz.interp_membership(value_range, identical_membership, value_diff)
            different_level = fuzz.interp_membership(value_range, different_membership, value_diff)

            # Defuzzification
            membership_values = np.column_stack((identical_level, different_level)).T
            similarity = fuzz.defuzz(value_range, membership_values, 'centroid')
            similarity_output = 'Identical' if similarity < 0 else 'Different'

            comparison_results.append({
                'row': row,
                'col': col,
                'value1': cell1_value,
                'value2': cell2_value,
                'similarity': similarity_output
            })

    return comparison_results

def main():
    st.title("Excel File Comparison using Fuzzy Logic")

    file1 = st.file_uploader("Upload Excel File 1", type=["xlsx"])
    file2 = st.file_uploader("Upload Excel File 2", type=["xlsx"])

    if file1 and file2:
        # Create in-memory file-like objects
        file1_bytes = file1.getvalue()
        file2_bytes = file2.getvalue()

        comparison_results = compare_excel_files(file1_bytes, file2_bytes)

        st.write("Comparison Results:")
        for result in comparison_results:
            st.write(f"Cell ({result['row']}, {result['col']}): Value 1 = {result['value1']}, Value 2 = {result['value2']}, Similarity = {result['similarity']}")

if __name__ == "__main__":
    main()